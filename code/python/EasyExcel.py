"""
简易使用excel
"""
import openpyxl as pyxl


class EasyExcel(object):
    """基础excel操作"""

    def __init__(self, excel_name: str):
        """ excel_name存在则打开，不存在则新建"""
        self.excel_name = excel_name
        self.sheet_name = "Sheet"
        self.row = 1
        try:
            self.excel = pyxl.load_workbook(excel_name)
        except FileNotFoundError:
            print(excel_name + "文件不存在,正在创建......")
            self.excel = pyxl.Workbook()
            self.excel.save(excel_name)
            self.excel = pyxl.load_workbook(excel_name)
            print("创建成功!")
        self.excel.active
        self.sheet = self.excel[self.excel.sheetnames[0]]

    def return_excel_name(self):
        return self.excel_name

    def return_excel_sheet_name(self):
        return self.sheet_name

    def create_sheet(self, new_sheet: str):
        self.excel.create_sheet(new_sheet)
        self.excel.save(self.excel_name)

    def rename_sheet(self, old_sheet: str, new_sheet_name: str):
        sheet = self.excel[old_sheet]
        sheet.title = new_sheet_name
        self.excel.save(self.excel_name)

    def set_sheet(self, sheet_name: str):
        self.sheet = self.excel[sheet_name]

    def save(self):
        self.excel.save(self.excel_name)


def move_all_col(excel1: EasyExcel, col1: int, excel2: EasyExcel, col2: int):
    for i in range(1, excel1.sheet.max_row+1):
        print(excel1.sheet.cell(row=i, column=col1).value)
        excel2.sheet.cell(row=i, column=col2).value = excel1.sheet.cell(
            row=i, column=col1).value
    excel2.save()


def move_part_col(excel1: EasyExcel, range1: list,
                  excel2: EasyExcel, range2: list):
    if len(range1) != len(range2):
        print("转移前后范围不同")
        return False
    pass


if __name__ == '__main__':
    x = EasyExcel("62.xlsx")
    y = EasyExcel("1.xlsx")
    x.set_sheet("2")
    move_all_col(y, 2, x, 3)
